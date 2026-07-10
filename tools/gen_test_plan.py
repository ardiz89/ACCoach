"""Generate src/accoach/web/test_plan.json from HONE_Test_Plan_AC_ACC.xlsx.

The XLSX stays the source of truth for the on-track test plan; the web tablet UI
reads the generated JSON. Re-run this whenever the plan changes:

    python tools/gen_test_plan.py
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "HONE_Test_Plan_AC_ACC.xlsx"
OUT = ROOT / "src" / "accoach" / "web" / "test_plan.json"

CATEGORIES = ["GT3", "Formula", "Stradali", "Generale (una volta)"]
CAT_ID = {"GT3": "GT3", "Formula": "Formula", "Stradali": "Stradali",
          "Generale (una volta)": "Generale"}


def _cell(v) -> str:
    return "" if v is None else str(v).strip()


def _header_row(rows) -> int:
    for i, r in enumerate(rows):
        if r and _cell(r[0]) == "#":
            return i
    raise ValueError("no header row (col0 == '#') found")


def _category(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    hdr = _header_row(rows)
    title = _cell(rows[0][0]) if rows else ws.title
    intro = _cell(rows[1][0]) if len(rows) > 1 else ""
    cid = CAT_ID[ws.title]
    tests = []
    for r in rows[hdr + 1:]:
        num = _cell(r[0])
        what = _cell(r[1]) if len(r) > 1 else ""
        if not num or not what:
            continue
        tests.append({
            "id": f"{cid}-{num}",
            "n": num,
            "what": what,
            "how": _cell(r[2]) if len(r) > 2 else "",
            "expected": _cell(r[3]) if len(r) > 3 else "",
        })
    return {"id": cid, "title": title, "intro": intro, "tests": tests}


def _glossary(wb) -> list[dict]:
    if "Glossario" not in wb.sheetnames:
        return []
    rows = list(wb["Glossario"].iter_rows(values_only=True))
    hdr = _header_row(rows)
    out = []
    for r in rows[hdr + 1:]:
        term = _cell(r[1]) if len(r) > 1 else ""
        meaning = _cell(r[2]) if len(r) > 2 else ""
        if term:
            out.append({"term": term, "meaning": meaning})
    return out


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    data = {
        "generated_from": XLSX.name,
        "categories": [_category(wb[name]) for name in CATEGORIES
                       if name in wb.sheetnames],
        "glossary": _glossary(wb),
    }
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    n = sum(len(c["tests"]) for c in data["categories"])
    print(f"wrote {OUT.relative_to(ROOT)}: {len(data['categories'])} categories, "
          f"{n} tests, {len(data['glossary'])} glossary terms")


if __name__ == "__main__":
    main()
